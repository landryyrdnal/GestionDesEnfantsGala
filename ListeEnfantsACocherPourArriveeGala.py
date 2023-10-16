import openpyxl
import unidecode

import data_base_process
import toml
import re
import datetime
import openpyxl
from openpyxl.styles import *
import unidecode


var_semaine, useless = data_base_process.fill_planning()
liste_profs = toml.load("parameters.toml")["profs"]
db = openpyxl.load_workbook("Gala 2024 ordre des cours.xlsx")
sheet = db.active


class CoursOrdreGala:
    def __init__(self, jour, heure, prof, duree, gala, ordre_passage):
        self.jour = jour
        self.heure = heure
        self.prof = prof
        self.duree = duree
        self.gala = gala
        self.ordre_passage = ordre_passage

    def __repr__(self):
        if self.ordre_passage < 10:
            ordre = "0" + str(self.ordre_passage)
        else:
            ordre = str(self.ordre_passage)
        return str(" G" + str(self.gala) + " T" + ordre + " " + self.jour[:2] + " " + self.heure + " " + self.prof)

    def __str__(self):
        return self.__repr__()

    def __eq__(self, other):
        if self.jour == other.jour and self.heure == other.heure and self.prof == other.prof and self.gala == other.gala and self.ordre_passage == other.ordre_passage:
            return True
        else:
            return False


liste_cours_horaires = []

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    cours = CoursOrdreGala(jour=row[1].value,
                           heure=row[2].value,
                           prof=row[3].value,
                           duree=row[8].value,
                           gala=row[9].value,
                           ordre_passage=row[10].value)
    liste_cours_horaires.append(cours)


# cherche dans la liste des cours avec durée le cours correspondant
def find_the_good_class_time(jour, heure, prof):
    for course in liste_cours_horaires:
        if course.heure == heure and course.jour == jour and course.prof == prof["nom"]:
            return course

    raise AssertionError


def search_heure(string):
    try:
        pattern = re.compile("[0-9][0-9]h[0-9][0-9]")
        return pattern.findall(string)[0]
    except IndexError:
        pattern = re.compile("[0-9]h[0-9][0-9]")
        return "0" + pattern.findall(string)[0]


def search_prof(string):
    for prof in liste_profs:
        if re.search(liste_profs[prof]["nom"], string):
            return liste_profs[prof]



def search_jour(string):
    jours = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
    for jour in jours:
        if re.search(jour, string):
            return jour


students = []
for level in var_semaine:
    for course in level:
        for student in course.liste_eleves:
            students.append(student)
print(len(students))
print(students)

student_list = []
for student in students:
    G1_courses = []
    G2_courses = []
    G3_courses = []
    principal_course = find_the_good_class_time(jour=student["cours"].jour,
                                                heure=student["cours"].heure,
                                                prof=student["cours"].prof)
    if principal_course.gala == 1 and principal_course not in G1_courses:
        G1_courses.append(principal_course)
    elif principal_course.gala == 2 and principal_course not in G2_courses:
        G2_courses.append(principal_course)
    elif principal_course.gala == 3 and principal_course not in G3_courses:
        G3_courses.append(principal_course)
    # on ajoute les autres cours
    for string in student["Autres cours"].split("|"):
        if len(string) > 1:
            prof = search_prof(string)
            heure = search_heure(string)
            jour = search_jour(string)
            course = find_the_good_class_time(jour, heure, prof)
            if course.gala == 1 and course not in G1_courses:
                G1_courses.append(course)
            elif course.gala == 2 and course not in G2_courses:
                G2_courses.append(course)
            elif course.gala == 3 and course not in G3_courses:
                G3_courses.append(course)
    eleve = {"nom": student["Nom"], "prénom": student["Prénom"], "G1": G1_courses, "G2": G2_courses,
             "G3": G3_courses, "téléphone": student["Téléphone"]}
    if eleve not in student_list:
        student_list.append(eleve)

print(student_list)
print(len(student_list))
student_list.sort(key=lambda x:  unidecode.unidecode(x["nom"]) and unidecode.unidecode(x["prénom"]))


# Génération du tableau
wb = openpyxl.Workbook()
# on enlève la feuille par défaut
wb.remove(wb['Sheet'])
# styles de mise en forme
#titre
st_titres = NamedStyle(name="style_titre")
st_titres.font = Font(name="Arial", size=14, bold=True)
st_titres.alignment = Alignment(horizontal="center", vertical="center")
#
g1 = NamedStyle(name="style_g1")
g1.font = Font(name="Arial", size=16, bold=False, color="FF0000")
g1.alignment = Alignment(horizontal="center", vertical="center")
g2 = NamedStyle(name="style_g2")
g2.font = Font(name="Arial", size=16, bold=False, color="FFFF00")
g2.alignment = Alignment(horizontal="center", vertical="center")
g3 = NamedStyle(name="style_g3")
g3.font = Font(name="Arial", size=16, bold=False)
g3.alignment = Alignment(horizontal="center", vertical="center")
st_temps = NamedStyle(name="style_temps")
st_temps.font = Font(name="Arial", size=11, bold=True)
st_temps.alignment = Alignment(horizontal="center", vertical="center")
st_normal = NamedStyle(name="style_normal")
st_normal.font = Font(name="Arial", size=11, bold=False)
st_normal.alignment = Alignment(horizontal="left", vertical="center")

for G in ["❶", "❷", "❸"]:
    ws = wb.create_sheet(G)
    ws.title = G
    ws.append([G, "", "Samedi 11 mars","","", "dimanche 12 mars"])
    ws[ws.max_row][0].style, ws[ws.max_row][1].style, ws[ws.max_row][2].style, ws[ws.max_row][3].style, ws[ws.max_row][4].style, ws[ws.max_row][5].style = st_titres, st_titres, st_titres, st_titres, st_titres, st_titres
    ws.append(["Prénom", "Nom", "Entrée", "Sortie", "Gala précédent", "Entrée", "Téléphone", "Cours"])
    ws[ws.max_row][0].style, ws[ws.max_row][1].style, ws[ws.max_row][2].style, ws[ws.max_row][3].style, ws[ws.max_row][4].style, ws[ws.max_row][5].style, ws[ws.max_row][6].style, ws[ws.max_row][7].style = st_titres, st_titres, st_titres, st_titres, st_titres, st_titres, st_titres, st_titres

    if G == "❶":
        already_in_list = []
        for student in student_list:
            if len(student["G1"]) >= 1:
                cours = ""
                for i in student["G1"]:
                    cours += str(i)
                    cours += ","
                cours = cours[:-1]
                # Condition pour éviter d’avoir un doublon
                if student["prénom"]+student["nom"] not in already_in_list:
                    #Ajout de la ligne de l’élève dans le tableau du gala 1
                    ws.append([student["prénom"], student["nom"].replace(" ","_"), "", "", "", "", student["téléphone"], cours])
                    already_in_list.append(student["prénom"]+student["nom"])
    if G == "❷":
        already_in_list = []
        for student in student_list:
            if len(student["G2"]) >= 1:
                cours = ""
                for i in student["G2"]:
                    cours += str(i)
                    cours += ","
                cours = cours[:-1]
                gala = ""
                # Condition pour éviter d’avoir un doublon
                if student["prénom"] + student["nom"] not in already_in_list:
                    if len(student["G1"]) >= 1:
                        gala = "❶"
                    ws.append([student["prénom"], student["nom"].replace(" ","_"), "", "", gala, "", student["téléphone"], cours])
                    already_in_list.append(student["prénom"] + student["nom"])
                    if len(student["G1"]) >= 1:
                        ws[ws.max_row][4].style = g1
    if G == "❸":
        already_in_list = []
        for student in student_list:
            if len(student["G3"]) >= 1:
                cours = ""
                for i in student["G3"]:
                    cours += str(i)
                    cours += ","
                cours = cours[:-1]
                gala = ""
                if len(student["G2"]) >= 1:
                    gala = "❷"
                if student["prénom"] + student["nom"] not in already_in_list:
                    ws.append([student["prénom"], student["nom"].replace(" ","_"), "", "", gala, "", student["téléphone"], cours])
                    already_in_list.append(student["prénom"] + student["nom"])
                    if len(student["G2"]) >= 1:
                        ws[ws.max_row][4].style = g2

    # largeur de colonnes
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10

wb.save("Tableau Enfants à cocher Gala.xlsx")
print("Tableau des enfants à cocher est sauvegardé")
