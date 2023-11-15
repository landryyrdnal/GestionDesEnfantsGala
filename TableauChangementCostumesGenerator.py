import openpyxl
import data_base_process
import toml
import re
import datetime
import openpyxl
from openpyxl.styles import *
import pandas as pd
import parameters


var_semaine, useless = data_base_process.process_data_base()
liste_profs = parameters.liste_profs

# Ouverture du fichier excel qui gère l’ordre de passage du gala de l’année
ordre_galas = parameters.ordre_galas
workbook = pd.read_excel(ordre_galas)



class CoursOrdreGala:
    def __init__(self, jour, heure, prof, duree, gala, ordre_passage):
        self.jour = jour
        self.heure = heure
        self.prof = prof
        self.duree = duree
        self.gala = gala
        self.ordre_passage = ordre_passage

    def __repr__(self):
        if self.ordre_passage < 10:
            ordre = "0" + str(self.ordre_passage)
        else:
            ordre = str(self.ordre_passage)
        return str(" G" + str(self.gala) + " T" + ordre + " " + self.jour[:2] + " " + self.heure + " " + self.prof)

    def __str__(self):
        return self.__repr__()

    def __eq__(self, other):
        if self.jour == other.jour and self.heure == other.heure and self.prof == other.prof and self.gala == other.gala and self.ordre_passage == other.ordre_passage:
            return True
        else:
            return False

# Analyse du fichier excel de l’ordre de passage du gala de l’année et création de la liste des cours
liste_cours_horaires = []

for index, row in workbook.iterrows():
    cours = CoursOrdreGala(jour=row["Jour"],
                           heure=row["Heure"],
                           prof=row["Professeur"],
                           duree=row["Durée"],
                           gala=row["Gala"],
                           ordre_passage=row["Ordre de Passage"])
    liste_cours_horaires.append(cours)


# cherche dans la liste des cours avec durée le cours correspondant
def find_the_good_class_time(jour, heure, prof):
    print(jour, heure, prof["nom"])
    for course in liste_cours_horaires:
        if course.heure == heure and course.jour == jour and course.prof == prof["nom"]:
            return course

    raise AssertionError


def search_heure(string):
    try:
        pattern = re.compile("[0-9][0-9]h[0-9][0-9]")
        return pattern.findall(string)[0]
    except IndexError:
        pattern = re.compile("[0-9]h[0-9][0-9]")
        return "0" + pattern.findall(string)[0]


def search_prof(string):
    for prof in liste_profs:
        if re.search(liste_profs[prof]["nom"], string):
            return liste_profs[prof]
        #elif re.search(liste_profs[prof]["diminutif"], string):
        #    return liste_profs[prof]


def search_jour(string):
    jours = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
    for jour in jours:
        if re.search(jour, string):
            return jour


students_with_multiple_courses = []
for level in var_semaine:
    for course in level:
        for student in course.liste_eleves:
            # si l'élève a d'autres cours on l'ajoute dans la liste des élèves ayant plusieurs cours
            if len(student["Autres cours"]) > 0:
                students_with_multiple_courses.append(student)

student_list = []

for student in students_with_multiple_courses:
    G1_courses = []
    G2_courses = []
    G3_courses = []
    # on ajoute son cours principal
    principal_course = find_the_good_class_time(jour=student["cours"].jour,
                                                heure=student["cours"].heure,
                                                prof=student["cours"].prof)
    if principal_course.gala == 1 and principal_course not in G1_courses:
        G1_courses.append(principal_course)
    elif principal_course.gala == 2 and principal_course not in G2_courses:
        G2_courses.append(principal_course)
    elif principal_course.gala == 3 and principal_course not in G3_courses:
        G3_courses.append(principal_course)
    # on ajoute les autres cours
    for string in student["Autres cours"].split("|"):
        if len(string) > 1:
            prof = search_prof(string)
            heure = search_heure(string)
            jour = search_jour(string)
            course = find_the_good_class_time(jour, heure, prof)
            if course.gala == 1 and course not in G1_courses:
                G1_courses.append(course)
            elif course.gala == 2 and course not in G2_courses:
                G2_courses.append(course)
            elif course.gala == 3 and course not in G3_courses:
                G3_courses.append(course)

            eleve = {"nom": student["Nom"], "prénom": student["Prénom"], "G1": G1_courses, "G2": G2_courses,
                     "G3": G3_courses}
            if eleve not in student_list:
                student_list.append(eleve)

student_dancing_at_least_twice_in_a_G = []
for student in student_list:
    if len(student["G1"]) >= 2 or len(student["G2"]) >= 2 or len(student["G3"]) >= 2:
        if len(student["G1"]) >= 2:
            student["G1"].sort(key=lambda x: x.ordre_passage)

        else:
            student["G1"] = []
        if len(student["G2"]) >= 2:
            student["G2"].sort(key=lambda x: x.ordre_passage)

        else:
            student["G2"] = []
        if len(student["G3"]) >= 2:
            student["G3"].sort(key=lambda x: x.ordre_passage)

        else:
            student["G3"] = []
        if student not in student_dancing_at_least_twice_in_a_G:
            student_dancing_at_least_twice_in_a_G.append(student)


def time_between_two_courses(first: CoursOrdreGala, second: CoursOrdreGala):
    # retourne le temps de changement entre deux passages pour se changer
    found = False
    index = liste_cours_horaires.index(first) + 1
    tot_time = datetime.timedelta()
    while not found:
        if liste_cours_horaires[index] != second:
            tot_time += datetime.timedelta(hours=liste_cours_horaires[index].duree.hour,
                                           minutes=liste_cours_horaires[index].duree.minute,
                                           seconds=liste_cours_horaires[index].duree.second)
            index += 1
        else:
            found = True
        minutes = tot_time.seconds // 60
        secondes = tot_time.seconds % 60
        if minutes >= 60:
            heures = minutes // 60
            minutes = minutes % 60
        else:
            heures = 00
    return str(heures) + ":" + str(minutes) + ":" + str(secondes)


for student in student_dancing_at_least_twice_in_a_G:
    galas = ["G1", "G2", "G3"]
    student["interval G1"] = []
    student["interval G2"] = []
    student["interval G3"] = []
    if len(student["G1"]) == 2:
        interval_G1 = time_between_two_courses(student["G1"][0], student["G1"][1])
        student["interval G1"].append(interval_G1)
    elif len(student["G1"]) == 3:
        interval_G1 = time_between_two_courses(student["G1"][0], student["G1"][1])
        interval2_G1 = time_between_two_courses(student["G1"][1], student["G1"][2])
        student["interval G1"].append(interval_G1)
        student["interval G1"].append(interval2_G1)
    if len(student["G2"]) == 2:
        interval_G2 = time_between_two_courses(student["G2"][0], student["G2"][1])
        student["interval G2"].append(interval_G2)
    elif len(student["G2"]) == 3:
        interval_G2 = time_between_two_courses(student["G2"][0], student["G2"][1])
        interval2_G2 = time_between_two_courses(student["G2"][1], student["G2"][2])
        student["interval G2"].append(interval_G2)
        student["interval G2"].append(interval2_G2)
    if len(student["G3"]) == 2:
        interval_G3 = time_between_two_courses(student["G3"][0], student["G3"][1])
        student["interval G3"].append(interval_G3)
    elif len(student["G3"]) == 3:
        interval_G3 = time_between_two_courses(student["G3"][0], student["G3"][1])
        interval2_G3 = time_between_two_courses(student["G3"][1], student["G3"][2])
        student["interval G3"].append(interval_G3)
        student["interval G3"].append(interval2_G3)

for i in student_dancing_at_least_twice_in_a_G:
    print(i)

# CRÉATION DU TABLEAU

# Génération du tableau
wb = openpyxl.Workbook()
# on enlève la feuille par défaut
wb.remove(wb['Sheet'])
# styles de mise en forme
st_titres = NamedStyle(name="style_titre")
st_titres.font = Font(name="Arial", size=14, bold=True)
st_titres.alignment = Alignment(horizontal="center", vertical="center")
st_titres_icon = NamedStyle(name="style_titre_icon")
st_titres_icon.font = Font(name="Arial", size=16, bold=False)
st_titres_icon.alignment = Alignment(horizontal="center", vertical="center")
st_temps = NamedStyle(name="style_temps")
st_temps.font = Font(name="Arial", size=11, bold=True)
st_temps.alignment = Alignment(horizontal="center", vertical="center")
st_normal = NamedStyle(name="style_normal")
st_normal.font = Font(name="Arial", size=11, bold=False)
st_normal.alignment = Alignment(horizontal="left", vertical="center")

for G in ["G1", "G2", "G3"]:
    ws = wb.create_sheet(G)
    ws.title = G
    ws.append(["Prénom", "Nom", "💃1", "⏱1", "💃2", "⏱2", "💃3"])
    ws[ws.max_row][0].style, ws[ws.max_row][1].style, ws[ws.max_row][2].style, ws[ws.max_row][3].style, ws[ws.max_row][
        4].style, ws[ws.max_row][5].style, ws[ws.max_row][
        6].style = st_titres, st_titres, st_titres_icon, st_titres_icon, st_titres_icon, st_titres_icon, st_titres_icon

    for student in student_dancing_at_least_twice_in_a_G:
        if len(student[G]) == 2:
            ws.append([student["prénom"], student["nom"], str(student[G][0]), student["interval " + G][0],
                       str(student[G][1])])
            ws[ws.max_row][0].style, ws[ws.max_row][1].style, ws[ws.max_row][2].style, ws[ws.max_row][3].style, \
            ws[ws.max_row][4].style, ws[ws.max_row][5].style, ws[ws.max_row][
                6].style = st_normal, st_normal, st_normal, st_temps, st_normal, st_temps, st_normal

        if len(student[G]) >= 3:
            ws.append(
                [student["prénom"], student["nom"], str(student[G][0]), student["interval " + G][0], str(student[G][1]),
                 str(student["interval " + G][1]), str(student[G][2])])
            ws[ws.max_row][0].style, ws[ws.max_row][1].style, ws[ws.max_row][2].style, ws[ws.max_row][3].style, \
            ws[ws.max_row][4].style, ws[ws.max_row][5].style, ws[ws.max_row][
                6].style = st_normal, st_normal, st_normal, st_temps, st_normal, st_temps, st_normal
    # largeur de colonnes
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 23
    ws.column_dimensions["C"].width = 31
    ws.column_dimensions["D"].width = 9
    ws.column_dimensions["E"].width = 31
    ws.column_dimensions["F"].width = 9
    ws.column_dimensions["G"].width = 31

wb.save("Tableau Changement de costumes.xlsx")
print("Tableau des changements de costume enregistré")
