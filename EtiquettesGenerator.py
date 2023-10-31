import data_base_process
from data_base_process import liste_couleurs
import unidecode
from openpyxl.worksheet import page
from openpyxl.worksheet.pagebreak import Break, RowBreak
from openpyxl import *
from openpyxl.styles import *
import gala_library

# GÉNÉRATION DU TABLEAU D'ÉTIQUETTES
# Génère le tableau Étiquettes.xlsx avec cinq pages pour toutes les étiquettes du Gala
#   première page :     Étiquette plastifiée à laisser sur le sac
#   deuxième page :     À donner par l'enfant en arrivant en répétition
#   troisième page :    À garder pour récupérer l’enfant après la répétition
#   quatrième page :    À donner par l'enfant en arrivant au spectacle
#   cinquième page :    Étiquette Costume Gala


var_semaine, liste_niveaux = data_base_process.fill_planning()
cours_gala = gala_library.ordre_de_passage_creator()


def construction_tableau_etiquettes():
    def str_cours(cours, tableau_gala):
        gala = "G" + str(tableau_gala.gala)
        if tableau_gala.ordre_passage < 10:
            tableau = "T0"+str(tableau_gala.ordre_passage)
        else:
            tableau = "T"+str(tableau_gala.ordre_passage)
        return gala + tableau + " " + cours.discipline[:4] + " " + cours.jour[:2] + \
        " " + cours.heure + " " + cours.prof["diminutif"]

    def eleve_etiquette(cours):
        """"
        retourne une liste d'élève dans laquelle chaque élève est un dictionnaire
        """
        # on cherche le tableau correspondant du cours dans la liste du gala pour indiquer
        # le n° du gala et le n° d’ordre de passage
        for tableau in cours_gala:
            if tableau.jour == cours.jour and tableau.heure == cours.heure and tableau.prof == cours.prof["nom"]:
                print(cours.jour, cours.heure, cours.prof["nom"], "|", tableau.jour, tableau.heure, tableau.prof)
                tableau_correspondant = tableau

        liste_eleve_cours = []
        # pour chaque élève on fait un dictionnaire qu'on ajoute dans la liste des élèves du cours
        for index, ligne in cours.df.iterrows():
            eleve = {}
            eleve["Cours"] = str_cours(cours, tableau_correspondant)
            eleve["Couleur"] = cours.prof["couleur_sac"]
            eleve["Couleur_style"] = f"style_couleur_{cours.prof['fond']}"
            eleve["Nom"] = ligne["Nom"]
            eleve["Prénom"] = ligne["Prénom"]
            eleve["Âge"] = ligne["Âge"]
            eleve["Autres cours"] = ligne["Autres cours"].split(" | ")
            # on renomme les autres cours : Co Ve 17h30 Nat
            for i in eleve["Autres cours"]:
                if i == "" or i == "nan":  # or math.isnan(i)
                    eleve["Autres cours"].remove(i)
                for niveau in liste_niveaux:
                    for coursname in niveau.liste_cours:
                        if coursname.nom_cours == i:
                            for tableau in cours_gala:
                                if tableau.jour == coursname.jour and tableau.heure == coursname.heure and tableau.prof == coursname.prof["nom"]:
                                    eleve["Autres cours"][eleve["Autres cours"].index(i)] = str_cours(coursname, tableau)
            # On enlève les cours dans Autres cours qui sont en double
            autres_cours_sans_doublon = []
            for _autre_cours in eleve["Autres cours"]:
                if _autre_cours not in autres_cours_sans_doublon:
                    autres_cours_sans_doublon.append(_autre_cours)
            eleve["Autres cours"] = autres_cours_sans_doublon
            # on tri les autres cours pour qu’ils apparaissent dans l’ordre des galas et des tableaux
            eleve["Autres cours"] = sorted(eleve["Autres cours"])
            liste_eleve_cours.append(eleve)
        liste_eleve_cours = index_eleves(liste_eleve_cours)
        # on retourne la liste
        return liste_eleve_cours

    def index_eleves(liste_eleve_cours):
        # on ajoute l'index de chaque élève
        liste_eleve_cours.sort(key=lambda x: unidecode.unidecode(x["Prénom"]))
        for i in liste_eleve_cours:
            i["Index"] = liste_eleve_cours.index(i) + 1
        # on retourne la liste
        return liste_eleve_cours

    # lister tous les élèves de chaque cours en prenant en compte les cours en doublons et
    # en les fusionnant pour qu'aucun élève ne soit oublié car il manque des élèves dans certains doublons…
    liste_index = []
    liste_eleves_cours = []
    for niveau in liste_niveaux:
        for cours in niveau.liste_cours:
            # on fabrique un identifiant pour le cours pour trouver les doublons
            code = cours.discipline + cours.prof["diminutif"] + cours.jour + cours.heure
            if code not in liste_index:
                liste_index.append(code)
                # on génère la liste des élèves dans le cours (un élève: dictionnaire)
                liste_eleves_cours.append(eleve_etiquette(cours))
            else:
                # on génère la liste des élèves dans le cours en doublon pour pouvoir ensuite la comparer avec la
                # liste originale
                doublon = eleve_etiquette(cours)
                # on compare la liste d'origine avec la liste doublon
                for eleve in liste_eleves_cours[liste_index.index(code)]:
                    for eleve_doublon in doublon:
                        # on retire de la liste des doublons les élèves qui se trouvent dans la liste originale
                        if eleve["Nom"] == eleve_doublon["Nom"] and eleve["Prénom"] == eleve_doublon["Prénom"]:
                            doublon.remove(eleve_doublon)
                # on ajoute les élèves restant s'il y en a à la liste originale de sorte à ce qu'il y ait tous les
                # élèves dans la liste originale
                for eleve_restant in doublon:
                    liste_eleves_cours[liste_index.index(code)].append(eleve_restant)
                    # on réactualise l'index des élèves
                    liste_eleves_cours[liste_index.index(code)] = index_eleves(
                        liste_eleves_cours[liste_index.index(code)])

    # Pour chaque élève on enlève le cours actuel dans la liste Autres cours
    for cours in liste_eleves_cours:
        for eleve in cours:
            for autre_cours in eleve["Autres cours"]:
                if autre_cours == eleve["Cours"]:
                    eleve["Autres cours"].remove(eleve["Cours"])

    # Construction du tableau
    wb = Workbook()
    # on enlève la feuille par défaut
    wb.remove(wb['Sheet'])

    # === STYLES ===
    # bordure
    bordure = Side(border_style="thick", color="000000")
    # prenom
    style_prenom = NamedStyle(name="style_prenom")
    style_prenom.font = Font(name="Arial", size=16, bold=True)
    style_prenom.alignment = Alignment(horizontal="right", vertical="center")
    # nom
    style_nom = NamedStyle(name="style_nom")
    style_nom.font = Font(name="Arial", size=12)
    style_nom.alignment = Alignment(horizontal="center", vertical="bottom")
    # titre
    style_titre = NamedStyle(name="style_titre")
    style_titre.font = Font(name="Arial", size=14, bold=True, underline="single")
    style_titre.alignment = Alignment(horizontal="center", vertical="bottom")
    # index
    style_index = NamedStyle(name="style_index")
    style_index.font = Font(name="Arial", size=14, bold=True)
    style_index.alignment = Alignment(horizontal="right", vertical="center", text_rotation=180)

    # autres cours
    style_autres_cours = NamedStyle(name="style_autres_cours")
    style_autres_cours.font = Font(name="Arial", size=10)
    # couleur / commentaire
    style_couleur = NamedStyle(name="style_couleur")
    style_couleur.font = Font(name="Arial", size=12, bold=True)
    style_couleur.alignment = Alignment(horizontal="right", vertical="center")
    style_couleur.fill = PatternFill("solid", start_color=liste_couleurs["grisClair"])
    # couleur noir
    style_couleur_noir = NamedStyle(name="style_couleur_noir")
    style_couleur_noir.font = Font(name="Arial", size=12, bold=True, color=liste_couleurs["blanc"])
    style_couleur_noir.alignment = Alignment(horizontal="right", vertical="center")
    style_couleur_noir.fill = PatternFill("solid", start_color=liste_couleurs["noir"])
    # couleur orange
    style_couleur_orange = NamedStyle(name="style_couleur_orange")
    style_couleur_orange.font = Font(name="Arial", size=12, bold=True, color=liste_couleurs["blanc"])
    style_couleur_orange.alignment = Alignment(horizontal="right", vertical="center")
    style_couleur_orange.fill = PatternFill("solid", start_color=liste_couleurs["orange"])
    # couleur jaune
    style_couleur_jaune = NamedStyle(name="style_couleur_jaune")
    style_couleur_jaune.font = Font(name="Arial", size=12, bold=True, color=liste_couleurs["noir"])
    style_couleur_jaune.alignment = Alignment(horizontal="right", vertical="center")
    style_couleur_jaune.fill = PatternFill("solid", start_color=liste_couleurs["jaune"])
    # couleur rose
    style_couleur_rose = NamedStyle(name="style_couleur_rose")
    style_couleur_rose.font = Font(name="Arial", size=12, bold=True, color=liste_couleurs["blanc"])
    style_couleur_rose.alignment = Alignment(horizontal="right", vertical="center")
    style_couleur_rose.fill = PatternFill("solid", start_color=liste_couleurs["rose"])
    # couleur rougeClair
    style_couleur_rougeClair = NamedStyle(name="style_couleur_rougeClair")
    style_couleur_rougeClair.font = Font(name="Arial", size=12, bold=True, color=liste_couleurs["blanc"])
    style_couleur_rougeClair.alignment = Alignment(horizontal="right", vertical="center")
    style_couleur_rougeClair.fill = PatternFill("solid", start_color=liste_couleurs["rougeClair"])
    # couleur rougeFonce
    style_couleur_rougeFonce = NamedStyle(name="style_couleur_rougeFonce")
    style_couleur_rougeFonce.font = Font(name="Arial", size=12, bold=True, color=liste_couleurs["blanc"])
    style_couleur_rougeFonce.alignment = Alignment(horizontal="right", vertical="center")
    style_couleur_rougeFonce.fill = PatternFill("solid", start_color=liste_couleurs["rougeFonce"])
    # couleur bleuCiel
    style_couleur_bleuCiel = NamedStyle(name="style_couleur_bleuCiel")
    style_couleur_bleuCiel.font = Font(name="Arial", size=12, bold=True, color=liste_couleurs["blanc"])
    style_couleur_bleuCiel.alignment = Alignment(horizontal="right", vertical="center")
    style_couleur_bleuCiel.fill = PatternFill("solid", start_color=liste_couleurs["bleuCiel"])
    # couleur bleuFonce
    style_couleur_bleuFonce = NamedStyle(name="style_couleur_bleuFonce")
    style_couleur_bleuFonce.font = Font(name="Arial", size=12, bold=True, color=liste_couleurs["blanc"])
    style_couleur_bleuFonce.alignment = Alignment(horizontal="right", vertical="center")
    style_couleur_bleuFonce.fill = PatternFill("solid", start_color=liste_couleurs["bleuFonce"])
    # couleur mauve
    style_couleur_mauve = NamedStyle(name="style_couleur_mauve")
    style_couleur_mauve.font = Font(name="Arial", size=12, bold=True, color=liste_couleurs["blanc"])
    style_couleur_mauve.alignment = Alignment(horizontal="right", vertical="center")
    style_couleur_mauve.fill = PatternFill("solid", start_color=liste_couleurs["mauve"])
    # couleur gris
    style_couleur_gris = NamedStyle(name="style_couleur_gris")
    style_couleur_gris.font = Font(name="Arial", size=12, bold=True, color=liste_couleurs["blanc"])
    style_couleur_gris.alignment = Alignment(horizontal="right", vertical="center")
    style_couleur_gris.fill = PatternFill("solid", start_color=liste_couleurs["gris"])
    # couleur grisClair
    style_couleur_grisClair = NamedStyle(name="style_couleur_grisClair")
    style_couleur_grisClair.font = Font(name="Arial", size=12, bold=True, color=liste_couleurs["blanc"])
    style_couleur_grisClair.alignment = Alignment(horizontal="right", vertical="center")
    style_couleur_grisClair.fill = PatternFill("solid", start_color=liste_couleurs["grisClair"])
    # couleur bleuNoir
    style_couleur_bleuNoir = NamedStyle(name="style_couleur_bleuNoir")
    style_couleur_bleuNoir.font = Font(name="Arial", size=12, bold=True, color=liste_couleurs["blanc"])
    style_couleur_bleuNoir.alignment = Alignment(horizontal="right", vertical="center")
    style_couleur_bleuNoir.fill = PatternFill("solid", start_color=liste_couleurs["bleuNoir"])
    # couleur vert
    style_couleur_vert = NamedStyle(name="style_couleur_vert")
    style_couleur_vert.font = Font(name="Arial", size=12, bold=True, color=liste_couleurs["blanc"])
    style_couleur_vert.alignment = Alignment(horizontal="right", vertical="center")
    style_couleur_vert.fill = PatternFill("solid", start_color=liste_couleurs["vert"])
    # couleur violet
    style_couleur_violet = NamedStyle(name="style_couleur_violet")
    style_couleur_violet.font = Font(name="Arial", size=12, bold=True, color=liste_couleurs["blanc"])
    style_couleur_violet.alignment = Alignment(horizontal="right", vertical="center")
    style_couleur_violet.fill = PatternFill("solid", start_color=liste_couleurs["violet"])
    # couleur blanc
    style_couleur_blanc = NamedStyle(name="style_couleur_blanc")
    style_couleur_blanc.font = Font(name="Arial", size=12, bold=True, color=liste_couleurs["noir"])
    style_couleur_blanc.alignment = Alignment(horizontal="right", vertical="center")
    style_couleur_blanc.fill = PatternFill("solid", start_color=liste_couleurs["blanc"])
    # couleur rougeRose
    style_couleur_rougeRose = NamedStyle(name="style_couleur_rougeRose")
    style_couleur_rougeRose.font = Font(name="Arial", size=12, bold=True, color=liste_couleurs["blanc"])
    style_couleur_rougeRose.alignment = Alignment(horizontal="right", vertical="center")
    style_couleur_rougeRose.fill = PatternFill("solid", start_color=liste_couleurs["rougeRose"])
    # couleur bleuCielClair
    style_couleur_bleuCielClair = NamedStyle(name="style_couleur_bleuCielClair")
    style_couleur_bleuCielClair.font = Font(name="Arial", size=12, bold=True, color=liste_couleurs["noir"])
    style_couleur_bleuCielClair.alignment = Alignment(horizontal="right", vertical="center")
    style_couleur_bleuCielClair.fill = PatternFill("solid", start_color=liste_couleurs["bleuCielClair"])
    # DICTIONNAIRE DES COULEURS 
    dict_color = {"style_couleur_blanc": style_couleur_blanc,
                  "style_couleur_noir": style_couleur_noir,
                  "style_couleur_orange": style_couleur_orange,
                  "style_couleur_jaune": style_couleur_jaune,
                  "style_couleur_rose": style_couleur_rose,
                  "style_couleur_rougeClair": style_couleur_rougeClair,
                  "style_couleur_rougeFonce": style_couleur_rougeFonce,
                  "style_couleur_bleuCiel": style_couleur_bleuCiel,
                  "style_couleur_bleuFonce": style_couleur_bleuFonce,
                  "style_couleur_mauve": style_couleur_mauve,
                  "style_couleur_gris": style_couleur_gris,
                  "style_couleur_grisClair": style_couleur_grisClair,
                  "style_couleur_bleuNoir": style_couleur_bleuNoir,
                  "style_couleur_vert": style_couleur_vert,
                  "style_couleur_violet": style_couleur_violet,
                  "style_couleur_rougeRose": style_couleur_rougeRose,
                  "style_couleur_bleuCielClair": style_couleur_rougeRose}
    # commentaire
    style_commentaire = NamedStyle(name="style_commentaire")
    style_commentaire.font = Font(name="Arial", size=11)
    style_commentaire.alignment = Alignment(horizontal="center")
    style_commentaire.fill = PatternFill("solid", start_color=liste_couleurs["grisClair"])
    # age
    style_age = NamedStyle(name="style_age")
    style_age.font = Font(name="Arial", size=12)
    style_age.alignment = Alignment(horizontal="right", vertical="center")

    eleve_vide = {"Cours": " ", "Couleur": " ", "Nom": " ", "Prénom": " ", "Âge": " ", "Autres cours": " ",
                  "Index": " ", "Couleur_style": "style_couleur_blanc"}

    # on tri les cours par jour et par heure
    days = {"Lu": 1, "Ma": 2, "Me": 3, "Je": 4, "Ve": 5, "Sa": 6, "Di": 7}
    liste_eleves_cours.sort(reverse=False, key=lambda x: (days[x[0]["Cours"][11:13]], x[0]["Cours"][14:19]))

    # on ajoute des élèves blancs dans chaque cours (3 dans les cours où le nb d'élève est impaire et 2 dans le
    # cours où le nb d'élève est pair de sorte à faire un espace vide entre chaque cours
    for cours in liste_eleves_cours:
        if len(cours) % 2 == 0:
            cours.append(eleve_vide)
            cours.append(eleve_vide)
        else:
            cours.append(eleve_vide)
            cours.append(eleve_vide)
            cours.append(eleve_vide)

    # on détermine les différents messages imprimés dans les différentes pages
    messages = ["Étiquette plastifiée à laisser sur le sac",
                "À donner par l'enfant en arrivant en répétition",
                "À garder pour récupérer l’enfant après la répétition",
                "À donner par l'enfant en arrivant au spectacle",
                "Étiquette Costume Gala"]
    messages_a_imprimer = [["Étiquette plastifiée à laisser", "sur le sac"],
                           ["À donner par l'enfant en arrivant", "en répétition"],
                           ["À garder pour récupérer l’enfant", "après la répétition"],
                           ["À donner par l'enfant en arrivant", "au spectacle"],
                           ["Étiquette Costume Gala", ""]]

    for message in messages:
        # on créé une nouvelle page en fonction du message à afficher sur les étiquettes
        print("début impression de la page {}.".format(str(messages.index(message) + 1)))
        ws = wb.create_sheet(str(messages.index(message) + 1))
        ws.title = str(messages.index(message) + 1)

        compteur = 0

        for cours in liste_eleves_cours:
            # on affiche deux élèves à la fois
            col = 1
            for eleve in cours:
                try:
                    eleve2 = cours[cours.index(eleve) + 1]
                except IndexError:
                    pass
                if col == 1:
                    if compteur == 6:
                        compteur = 0

                    # nom prénom
                    ws.append([eleve["Nom"], " ", eleve["Prénom"], " ", eleve2["Nom"], " ", eleve2["Prénom"]])

                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=1, end_column=2)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=3, end_column=4)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=5, end_column=6)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=7, end_column=8)
                    ws[ws.max_row][0].style, ws[ws.max_row][4].style = style_nom, style_nom
                    ws[ws.max_row][2].style, ws[ws.max_row][6].style = style_prenom, style_prenom

                    # Cours
                    ws.append([eleve["Cours"], " ", " ", str(eleve["Index"]), eleve2["Cours"], " ", " ",
                               str(eleve2["Index"])])
                    ws[ws.max_row][0].style, ws[ws.max_row][4].style = style_titre, style_titre
                    ws[ws.max_row][3].style, ws[ws.max_row][7].style = style_index, style_index
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=1, end_column=3)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=5, end_column=7)
                    # Index / age
                    ws.append(["  " + " ", str(eleve["Âge"]) + " ans", " ", "•",
                               "  " + " ", str(eleve2["Âge"]) + " ans", " ", "•"])
                    # on redimensionne la hauteur de ligne
                    rd = ws.row_dimensions[ws.max_row]
                    rd.height = 23
                    ws[ws.max_row][1].style, ws[ws.max_row][5].style = style_age, style_age

                    # Autres cours
                    # premier rang autres cours
                    impr = []
                    espace = " "
                    if len(eleve["Autres cours"]) == 0:
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                    elif len(eleve["Autres cours"]) == 1:
                        impr.append(espace + eleve["Autres cours"][0])
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                    elif len(eleve["Autres cours"]) == 2:
                        impr.append(espace + eleve["Autres cours"][0])
                        impr.append(" ")
                        impr.append(eleve["Autres cours"][1])
                        impr.append(" ")
                    else:
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                    if len(eleve2["Autres cours"]) == 0:
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                    elif len(eleve2["Autres cours"]) == 1:
                        impr.append(espace + eleve2["Autres cours"][0])
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                    elif len(eleve2["Autres cours"]) == 2:
                        impr.append(espace + eleve2["Autres cours"][0])
                        impr.append(" ")
                        impr.append(eleve2["Autres cours"][1])
                        impr.append(" ")
                    else:
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                    ws.append(impr)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=1, end_column=2)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=3, end_column=4)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=5, end_column=6)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=7, end_column=8)
                    ws[ws.max_row][0].style, ws[ws.max_row][2].style = style_autres_cours, style_autres_cours
                    ws[ws.max_row][4].style, ws[ws.max_row][6].style = style_autres_cours, style_autres_cours

                    # deuxième rang autres cours
                    impr = []
                    if len(eleve["Autres cours"]) == 3:
                        impr.append(eleve["Autres cours"][2])
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                    elif len(eleve["Autres cours"]) >= 4:
                        impr.append(eleve["Autres cours"][2])
                        impr.append(" ")
                        impr.append(eleve["Autres cours"][3])
                        impr.append(" ")
                    else:
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                    if len(eleve2["Autres cours"]) == 3:
                        impr.append(eleve2["Autres cours"][2])
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                    elif len(eleve2["Autres cours"]) >= 4:
                        impr.append(eleve2["Autres cours"][2])
                        impr.append(" ")
                        impr.append(eleve2["Autres cours"][3])
                        impr.append(" ")
                    else:
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                    ws.append(impr)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=1, end_column=2)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=3, end_column=4)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=5, end_column=6)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=7, end_column=8)
                    ws[ws.max_row][0].style, ws[ws.max_row][2].style = style_autres_cours, style_autres_cours
                    ws[ws.max_row][4].style, ws[ws.max_row][6].style = style_autres_cours, style_autres_cours

                    # couleur + message
                    ws.append([messages_a_imprimer[messages.index(message)][0], " ", " ", eleve["Couleur"],
                               messages_a_imprimer[messages.index(message)][0], " ", " ", eleve2["Couleur"]])
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=1, end_column=3)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=5, end_column=7)
                    ws[ws.max_row][0].style, ws[ws.max_row][4].style = style_commentaire, style_commentaire
                    ws[ws.max_row][3].style, ws[ws.max_row][7].style = dict_color[eleve["Couleur_style"]], \
                                                                       dict_color[eleve2["Couleur_style"]]
                    # deuxième ligne commentaire
                    ws.append([messages_a_imprimer[messages.index(message)][1], " ", " ", " ",
                               messages_a_imprimer[messages.index(message)][1], " ", " "])
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=1, end_column=3)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=5, end_column=7)
                    ws[ws.max_row][0].style, ws[ws.max_row][4].style = style_commentaire, style_commentaire
                    ws[ws.max_row][3].style, ws[ws.max_row][7].style = style_commentaire, style_commentaire
                    col = 2
                    compteur += 1
                else:
                    col = 1

        # ajout des bordures
        # verticales
        for row in ws:
            row[0].border = Border(left=bordure)
            row[3].border = Border(right=bordure)
            row[4].border = Border(left=bordure)
            row[7].border = Border(right=bordure)
        # horizontales
        compteur = 0
        for row in ws:
            compteur += 1
            if compteur == 1:
                row[0].border = Border(top=bordure, left=bordure)
                row[1].border = Border(top=bordure)
                row[2].border = Border(top=bordure)
                row[3].border = Border(top=bordure, right=bordure)
                row[4].border = Border(top=bordure, left=bordure)
                row[5].border = Border(top=bordure)
                row[6].border = Border(top=bordure)
                row[7].border = Border(top=bordure, right=bordure)
            elif compteur == 7:
                row[0].border = Border(bottom=bordure, left=bordure)
                row[1].border = Border(bottom=bordure)
                row[2].border = Border(bottom=bordure)
                row[3].border = Border(bottom=bordure, right=bordure)
                row[4].border = Border(bottom=bordure, left=bordure)
                row[5].border = Border(bottom=bordure)
                row[6].border = Border(bottom=bordure)
                row[7].border = Border(bottom=bordure, right=bordure)
                compteur = 0
        # définition de la larguer des colonnes
        ws.column_dimensions["A"].width = 11.90
        ws.column_dimensions["B"].width = 11.90
        ws.column_dimensions["C"].width = 11.90
        ws.column_dimensions["D"].width = 11.90
        ws.column_dimensions["E"].width = 11.90
        ws.column_dimensions["F"].width = 11.90
        ws.column_dimensions["G"].width = 11.90
        ws.column_dimensions["H"].width = 11.90

        # définition des sauts de page
        max_row = ws.max_row
        break_number = 0
        row_break = RowBreak()
        while break_number <= max_row:
            break_number += 42
            row_break.append(Break(id=break_number))
        ws.row_breaks = row_break

        # définition de la zone d'impression et des marges
        ws.page_margins = page.PageMargins(left=0.23, right=0.23, top=0.23, bottom=0.23, header=0.23, footer=0.23)
        wb.page_margins = page.PageMargins(left=0.23, right=0.23, top=0.23, bottom=0.23, header=0.23, footer=0.23)
        print("Impression de la page {} terminée".format(str(messages.index(message) + 1)))
    # sauvegarde du fichier
    wb.save("Étiquettes.xlsx")
    print("Le fichier a été sauvergardé sous le nom Étiquettes.xlsx")


if __name__ == "__main__":
    construction_tableau_etiquettes()
