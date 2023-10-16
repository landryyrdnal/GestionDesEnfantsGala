from data_base_process import liste_couleurs, liste_profs, semaine, fill_planning
from openpyxl import *
from openpyxl.styles import *
from openpyxl.utils.dataframe import dataframe_to_rows

# GÉNÉRATION DU TABLEAU D'APPEL
# Génère le fichier liste_appel.xlsx qui sert à faire des listes d’appel
var_semaine, useless = fill_planning()

def contsruction_tableau_appel():
    # écriture sur le fichier excel
    print("Début de la construction du tableau d’appel")
    wb = Workbook()
    ws = wb.active
    ws.title = "Lundi"
    wb.create_sheet("Mardi")
    wb.create_sheet("Mercredi")
    wb.create_sheet("Jeudi")
    wb.create_sheet("Vendredi")
    wb.create_sheet("Samedi")
    wb.create_sheet("Dimanche")

    # définition des styles
    row1 = NamedStyle(name="row1")
    row1.font = Font(name="Arial", size=10)
    row1.fill = PatternFill("solid", start_color=liste_couleurs["grisClair"])

    row2 = NamedStyle(name="row2")
    row2.font = Font(name="Arial", size=10)
    row2.fill = PatternFill("solid", start_color=liste_couleurs["blanc"])

    sub = NamedStyle(name="sub")
    sub.font = Font(name="Arial", size=11, color=liste_couleurs["blanc"])
    sub.fill = PatternFill("solid", start_color=liste_couleurs["gris"])

    prof_style = NamedStyle(name="prof")

    def mise_en_forme(cell, style):
        # applique le style à la cellule
        cell.style = style

    def mise_en_forme_prof(cell, prof):
        # change la couleur en fonction du prof
        if prof is None:
            prof = liste_profs["none"]
        prof_style.font = Font(name="Arial", size=13, color=liste_couleurs[prof["couleur"]])
        prof_style.fill = PatternFill("solid", start_color=liste_couleurs[prof["fond"]])
        mise_en_forme(cell, prof_style)

    for jour in semaine:
        ws = wb[jour]
        for cours in var_semaine[semaine.index(jour)]:
            if cours.prof is None:
                print("Impossible de déterminer la·le prof du cours : {}".format(cours.nom_cours))
            # titre du tableau
            ws.append([cours.nom_cours, "", "", "", "", cours.discipline, "Total élèves : " + str(cours.nb_eleves)])
            ws.merge_cells(start_row=ws.max_row, end_row=ws.max_row, start_column=1, end_column=5)
            ws.merge_cells(start_row=ws.max_row, end_row=ws.max_row, start_column=7, end_column=8)
            for cell in ws[ws.max_row]:
                mise_en_forme_prof(cell, cours.prof)

            # corps du tableau avec entêtes
            compteur = 1
            for row in dataframe_to_rows(cours.df, index=False, header=True):
                ws.append(row)
                # mise en forme

                for cell in ws[ws.max_row]:
                    if compteur == 1:
                        mise_en_forme(cell, sub)
                    elif compteur == 2:
                        mise_en_forme(cell, row2)
                    elif compteur == 3:
                        mise_en_forme(cell, row1)
                if compteur == 1:
                    compteur = 2
                elif compteur == 2:
                    compteur = 3
                elif compteur == 3:
                    compteur = 2
            ws.append([" ", " ", " ", " ", " "])
    for sheet in wb:
        sheet.column_dimensions["A"].width = 15
        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 5
        sheet.column_dimensions["D"].width = 8
        sheet.column_dimensions["E"].width = 25
        sheet.column_dimensions["F"].width = 15
        sheet.column_dimensions["G"].width = 19
        sheet.column_dimensions["H"].width = 8

    # sauvegarde du fichier xlsx de sortie
    print("Génération du tableau d'appel terminée dans liste_appel.xlsx")
    wb.save("liste_appel.xlsx")
