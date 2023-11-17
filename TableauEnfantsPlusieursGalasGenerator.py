import data_base_process
import toml
import re
import openpyxl
import unidecode
from openpyxl.styles import *
from parameters import liste_profs, ordre_galas

var_semaine, useless = data_base_process.process_data_base()
db = ordre_galas
sheet = db.active


class CoursOrdreGala:
    def __init__(self, jour, heure, prof, duree, gala, ordre_passage):
        self.jour = jour
        self.heure = heure
        self.prof = prof
        self.duree = duree
        self.gala = gala
        self.ordre_passage = ordre_passage

    def __repr__(self):
        if self.ordre_passage < 10:
            ordre = "0" + str(self.ordre_passage)
        else:
            ordre = str(self.ordre_passage)
        return str(" G" + str(self.gala) + " T" + ordre + " " + self.jour[:2] + " " + self.heure + " " + self.prof)

    def __str__(self):
        return self.__repr__()

    def __eq__(self, other):
        if self.jour == other.jour and self.heure == other.heure and self.prof == other.prof and self.gala == other.gala and self.ordre_passage == other.ordre_passage:
            return True
        else:
            return False


liste_cours_horaires = []

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    cours = CoursOrdreGala(jour=row[1].value,
                           heure=row[2].value,
                           prof=row[3].value,
                           duree=row[8].value,
                           gala=row[9].value,
                           ordre_passage=row[10].value)
    liste_cours_horaires.append(cours)


# cherche dans la liste des cours avec durée le cours correspondant
def find_the_good_class_time(jour, heure, prof):
    print(jour, heure, prof)
    for course in liste_cours_horaires:
        if course.heure == heure and course.jour == jour and course.prof == prof["nom"]:
            return course

    raise AssertionError


def search_heure(string):
    try:
        pattern = re.compile("[0-9][0-9]h[0-9][0-9]")
        return pattern.findall(string)[0]
    except IndexError:
        pattern = re.compile("[0-9]h[0-9][0-9]")
        return "0" + pattern.findall(string)[0]


def search_prof(string):
    for prof in liste_profs:
        if re.search(liste_profs[prof]["nom"], string):
            return liste_profs[prof]
        elif re.search(liste_profs[prof]["diminutif"], string):
            return liste_profs[prof]


def search_jour(string):
    jours = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
    for jour in jours:
        if re.search(jour, string):
            return jour


students = []
for level in var_semaine:
    for course in level:
        for student in course.liste_eleves:
            students.append(student)

student_list = []
for student in students:
    G1_courses = []
    G2_courses = []
    G3_courses = []
    principal_course = find_the_good_class_time(jour=student["cours"].jour,
                                                heure=student["cours"].heure,
                                                prof=student["cours"].prof)
    if principal_course.gala == 1 and principal_course not in G1_courses:
        G1_courses.append(principal_course)
    elif principal_course.gala == 2 and principal_course not in G2_courses:
        G2_courses.append(principal_course)
    elif principal_course.gala == 3 and principal_course not in G3_courses:
        G3_courses.append(principal_course)
    # on ajoute les autres cours
    for string in student["Autres cours"].split("|"):
        if len(string) > 1:
            prof = search_prof(string)
            heure = search_heure(string)
            jour = search_jour(string)
            course = find_the_good_class_time(jour, heure, prof)
            if course.gala == 1 and course not in G1_courses:
                G1_courses.append(course)
            elif course.gala == 2 and course not in G2_courses:
                G2_courses.append(course)
            elif course.gala == 3 and course not in G3_courses:
                G3_courses.append(course)
    eleve = {"nom": student["Nom"], "prénom": student["Prénom"], "G1": G1_courses, "G2": G2_courses,
             "G3": G3_courses, "téléphone": student["Téléphone"], "mail":student["Mail"]}
    if eleve not in student_list:
        student_list.append(eleve)

print(student_list)
print(len(student_list))
student_list.sort(key=lambda x: unidecode.unidecode(x["nom"]) and unidecode.unidecode(x["prénom"]))
st_titres = NamedStyle(name="style_titre")
st_titres.font = Font(name="Arial", size=14, bold=True)
st_titres.alignment = Alignment(horizontal="center", vertical="center")
st_titres_icon = NamedStyle(name="style_titre_icon")
st_titres_icon.font = Font(name="Arial", size=16, bold=False)
st_titres_icon.alignment = Alignment(horizontal="center", vertical="center")

# ce code n'est utile qu'en exploitant la sortie dans la console
for G in ["❶❷", "❷❸", "❶❸", "❶❷❸"]:
    ws = db.create_sheet(G)
    ws.title = G

    ws.append([G, "", "Samedi 16 mars", "", "", "dimanche 17 mars"])
    ws[ws.max_row][0].style, ws[ws.max_row][1].style, ws[ws.max_row][2].style, ws[ws.max_row][3].style, ws[ws.max_row][
        4].style, ws[ws.max_row][5].style = st_titres, st_titres, st_titres, st_titres, st_titres, st_titres
    ws.append(["Prénom", "Nom", "Entrée", "Sortie", "Gala précédent", "Entrée", "Téléphone", "Mail"])
    ws[ws.max_row][0].style, ws[ws.max_row][1].style, ws[ws.max_row][2].style, ws[ws.max_row][3].style, ws[ws.max_row][
        4].style, ws[ws.max_row][5].style, ws[ws.max_row][6].style, ws[ws.max_row][
        7].style = st_titres, st_titres, st_titres, st_titres, st_titres, st_titres, st_titres, st_titres

    if G == "❶❷":
        for student in student_list:
            if len(student["G1"]) >= 1 and len(student["G2"]) >= 1:
                print(student["prénom"] + "," + student["nom"] + "," + str(student["G1"]) + "," + str(student["G2"]))
                ws.append([student["prénom"], student["nom"], "", "", "❶", "", student["téléphone"], student["mail"]])
    if G == "❷❸":
        for student in student_list:
            if len(student["G2"]) >= 1 and len(student["G3"]) >= 1:
                print(student["prénom"] + "," + student["nom"] + "," + str(student["G2"]) + "," + str(student["G3"]))
                ws.append([student["prénom"], student["nom"], "", "", "❷", "", student["téléphone"], student["mail"]])
    if G == "❶❸":
        for student in student_list:
            if len(student["G1"]) >= 1 and len(student["G3"]) >= 1:
                print(student["prénom"] + "," + student["nom"] + "," + str(student["G1"]) + "," + str(student["G3"]))
                ws.append([student["prénom"], student["nom"], "", "", "❶", "", student["téléphone"], student["mail"]])
    if G == "❶❷❸":
        for student in student_list:
            if len(student["G1"]) >= 1 and len(student["G2"]) >= 1 and len(student["G3"]) >= 1:
                print(student["prénom"] + "," + student["nom"] + "," + str(student["G1"]) + "," + str(student["G3"]))
                ws.append([student["prénom"], student["nom"], "", "", "❶", "", student["téléphone"], student["mail"]])

db.save("Tableau Enfants dans plusieurs galas.xlsx")
print("Tableau Enfants dans plusieurs galas.xlsx enregistré")
